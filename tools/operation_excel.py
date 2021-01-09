"""
    @file: operation_excel.py
    @Author: Shieh
    @Date: 2021/1/2 21:59
    @Description:
"""
import json
import time

import openpyxl
from config import base_path, cell_config
import os


class OperationExcel:
    def __init__(self, filename):
        # 动态文件路径
        self.filename = base_path + os.sep + "data" + os.sep + filename
        # 打开文件
        self.workbook = openpyxl.load_workbook(self.filename)
        # 获取sheet表单对象
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 获取总行
        self.row = self.sheet.max_row
        print(self.row)

    def read_excel(self):
        case = list()
        # 遍历每行数据，存入字典
        for i in range(2, self.row + 1):
            data = dict()
            if self.sheet.cell(i, cell_config.get("is_run")).value == "Yes":
                try:
                    data['path'] = self.sheet.cell(i, cell_config.get("path")).value
                    data['method'] = str(self.sheet.cell(i, cell_config.get("method")).value).lower()
                    data['headers'] = eval(self.sheet.cell(i, cell_config.get("headers")).value)
                    data['param_type'] = self.sheet.cell(i, cell_config.get("param_type")).value
                    data['params'] = eval(self.sheet.cell(i, cell_config.get("params")).value)
                    data['expect'] = eval(self.sheet.cell(i, cell_config.get("expect")).value)
                    # 记录每行数据的行与列，执行完成写入结果使用
                    data['x_y'] = [i, cell_config.get("result")]
                    # 将字典追加到列表中
                    case.append(data)
                    # 将读取数据写入Excel
                    now_time = time.strftime('%Y-%m-%d %H:%M:%S')
                    self.write_excel([i, cell_config.get("desc")], "读取完成~!{}".format(now_time))
                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], e)
            # 将列表数据写入json
            self.write_json(case, "api.json")
        print(case)

    def write_excel(self, x_y, msg):
        """
        :param x_y: 参数为列表，如：[2, 5]  2行5列
        :param msg:
        :return:
        """
        try:
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 不管是否成功都要保存文件
            self.workbook.save(self.filename)

    @staticmethod
    def read_json(file_name="api.json"):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "r", encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def write_json(case, file_name):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "w", encoding="utf-8") as f:
            json.dump(case, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    OperationExcel("api.xlsx").read_excel()
